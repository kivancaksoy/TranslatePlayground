import json
import random
import time

import openpyxl
from deep_translator import GoogleTranslator, DeeplTranslator

from src.model.translation_dto import TranslationDto


def translate_text(text, row_index, col_index, translation_dto_file):
    try:
        if translation_dto_file.translate_service == "google":
            return GoogleTranslator(source=translation_dto_file.source_lang,
                                    target=translation_dto_file.target_lang).translate(text)
        elif translation_dto_file.translate_service == "deepl":
            return DeeplTranslator(api_key=translation_dto_file.deepl_api_key,
                                   source=translation_dto_file.source_lang,
                                   target=translation_dto_file.target_lang).translate(text)
        else:
            raise ValueError("Invalid translation service selected.")
    except Exception as e:
        print(f"Translation error at row {row_index}, column {col_index}: '{text}' -> {e}")
        return text


def translate_excel(translation_dto_file_path):
    translation_dto_file = load_dto_from_json(translation_dto_file_path)
    workbook = openpyxl.load_workbook(translation_dto_file.input_file)
    sheet = workbook.active

    translated_workbook = openpyxl.Workbook()
    translated_sheet = translated_workbook.active

    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        translated_row = []
        for col_index, cell in enumerate(row, start=1):
            if translation_dto_file.all_columns or (col_index in translation_dto_file.columns_to_translate):
                if isinstance(cell, str) and cell.strip():
                    translated = translate_text(cell, row_index, col_index, translation_dto_file)
                    translated_row.append(translated)
                    time.sleep(random.uniform(0.2, 0.6))
                else:
                    translated_row.append(cell)
            else:
                translated_row.append(cell)
        translated_sheet.append(translated_row)

    translated_workbook.save(translation_dto_file.output_file)
    print(f"Translated file saved as: {translation_dto_file.output_file}")


def load_dto_from_json(json_file):
    with open(json_file, "r") as file:
        data = json.load(file)
    return TranslationDto(
        input_file=data["input_file"],
        output_file=data["output_file"],
        translate_service=data["translate_service"],
        deepl_api_key=data["deepl_api_key"],
        source_lang=data["source_lang"],
        target_lang=data["target_lang"],
        columns_to_translate=data["columns_to_translate"],
        all_columns=data["all_columns"]
    )
