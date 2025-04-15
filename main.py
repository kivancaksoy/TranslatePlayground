import random

import openpyxl
from deep_translator import GoogleTranslator
import time


def translate_text(text, row_index, col_index):
    try:
        return GoogleTranslator(source='de', target='tr').translate(text)
    except Exception as e:
        print(f"Translation error at row {row_index}, column {col_index}: '{text}' -> {e}")
        return text


def translate_excel(input_file, output_file, columns_to_translate):
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active

    translated_workbook = openpyxl.Workbook()
    translated_sheet = translated_workbook.active

    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        translated_row = []
        for col_index, cell in enumerate(row, start=1):
            if col_index in columns_to_translate:
                if isinstance(cell, str) and cell.strip():
                    translated = translate_text(cell, row_index, col_index)
                    translated_row.append(translated)
                    time.sleep(random.uniform(0.2, 0.6))
                else:
                    translated_row.append(cell)
            else:
                translated_row.append(cell)
        translated_sheet.append(translated_row)

    translated_workbook.save(output_file)
    print(f"Translated file saved as: {output_file}")


if __name__ == '__main__':
    columns_to_translate = [3, 4, 5, 51]
    translate_excel("deneme.xlsx", "translated_file.xlsx", columns_to_translate)
